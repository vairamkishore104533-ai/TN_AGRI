from models.expense import Expense
from services.ai_service import AIService
from datetime import datetime, timedelta
import calendar


class ExpenseAnalytics:
    @staticmethod
    def get_chart_data(user_id):
        summary = Expense.get_summary(user_id)
        current_year = datetime.utcnow().year
        monthly = Expense.get_monthly_data(user_id, current_year)
        expense_breakdown = Expense.get_category_breakdown(user_id, "expense")
        income_breakdown = Expense.get_category_breakdown(user_id, "income")
        recent = Expense.get_recent(user_id, 10)
        stats = Expense.get_stats(user_id)
        total_income = summary.get("income", 0)
        total_expense = summary.get("expense", 0)
        net_profit = total_income - total_expense
        profit_margin = round((net_profit / total_income) * 100, 1) if total_income > 0 else 0
        savings_rate = round(((total_income - total_expense) / total_income) * 100, 1) if total_income > 0 else 0
        months_list = []
        for m in range(1, 13):
            months_list.append({
                "month": m,
                "month_name": calendar.month_abbr[m],
                "income": monthly.get(m, {}).get("income", 0),
                "expense": monthly.get(m, {}).get("expense", 0),
                "profit": monthly.get(m, {}).get("income", 0) - monthly.get(m, {}).get("expense", 0),
            })
        return {
            "summary": summary,
            "net_profit": net_profit,
            "profit_margin": profit_margin,
            "savings_rate": savings_rate,
            "stats": stats,
            "monthly": months_list,
            "expense_breakdown": expense_breakdown,
            "income_breakdown": income_breakdown,
            "recent": [e.to_dict() for e in recent],
        }

    @staticmethod
    def get_ai_insights(user_id, lang="en"):
        summary = Expense.get_summary(user_id)
        stats = Expense.get_stats(user_id)
        expense_breakdown = Expense.get_category_breakdown(user_id, "expense")
        income_breakdown = Expense.get_category_breakdown(user_id, "income")
        total_income = summary.get("income", 0)
        total_expense = summary.get("expense", 0)
        net = total_income - total_expense
        top_exp = expense_breakdown[:3] if expense_breakdown else []
        top_inc = income_breakdown[:3] if income_breakdown else []
        if not total_income and not total_expense:
            if lang == "ta":
                return "இதுவரை எந்த பரிவர்த்தனையும் பதிவு செய்யப்படவில்லை. உங்கள் வருமானம் மற்றும் செலவுகளை பதிவு செய்ய தொடங்குங்கள்."
            else:
                return "No transactions recorded yet. Start recording your income and expenses to get insights."

        if lang == "ta":
            prompt = (
                f"நீங்கள் தமிழ்நாடு விவசாய நிதி ஆலோசகர். பின்வரும் தரவுகளின் அடிப்படையில் சுருக்கமான நிதி ஆலோசனைகளை வழங்கவும்.\n\n"
                f"மொத்த வருமானம்: ₹{total_income}\n"
                f"மொத்த செலவு: ₹{total_expense}\n"
                f"நிகர லாபம்: ₹{net}\n"
                f"அதிக செலவு வகைகள்: "
            )
            for c in top_exp:
                prompt += f"{c['_id']}: ₹{c['total']}, "
            prompt += f"\nஅதிக வருமான ஆதாரங்கள்: "
            for c in top_inc:
                prompt += f"{c['_id']}: ₹{c['total']}, "
            prompt += (
                f"\n\nசுருக்கமான நிதி பகுப்பாய்வு மற்றும் 3 முக்கிய பரிந்துரைகளை தமிழில் வழங்கவும். "
                f"எடுத்துக்காட்டு: உர செலவுகளை குறைக்க பரிந்துரை, தொழிலாளர் செலவு கட்டுப்பாடு, நீர்ப்பாசன திறன் மேம்பாடு போன்றவை. "
                f"3-4 வரிகளுக்கு மேல் இருக்க வேண்டாம்."
            )
        else:
            prompt = (
                f"You are a Tamil Nadu farm financial advisor. Analyze the following farm financial data and provide concise insights.\n\n"
                f"Total Income: ₹{total_income}\n"
                f"Total Expenses: ₹{total_expense}\n"
                f"Net Profit: ₹{net}\n"
                f"Top Expense Categories: "
            )
            for c in top_exp:
                prompt += f"{c['_id']}: ₹{c['total']}, "
            prompt += f"\nTop Income Sources: "
            for c in top_inc:
                prompt += f"{c['_id']}: ₹{c['total']}, "
            prompt += (
                f"\n\nProvide a brief financial analysis and 3 key recommendations. "
                f"Examples: reduce fertilizer costs, control labor expenses, improve irrigation efficiency. "
                f"Keep it to 3-4 lines."
            )
        ai = AIService()
        try:
            return ai.get_response(prompt, lang)
        except Exception:
            if lang == "ta":
                return "உங்கள் நிதி நிலை சீராக உள்ளது. செலவுகளை கண்காணித்து வருமானத்தை அதிகரிக்க முயற்சிக்கவும்."
            else:
                return "Your financial position is stable. Continue monitoring expenses and look for ways to increase income."


class BudgetManager:
    @staticmethod
    def get_budgets(user_id):
        db = Expense.get_collection().database
        budgets_coll = db["expense_budgets"]
        docs = list(budgets_coll.find({"user_id": user_id}))
        for d in docs:
            d["_id"] = str(d["_id"])
        return docs

    @staticmethod
    def save_budgets(user_id, budgets):
        db = Expense.get_collection().database
        budgets_coll = db["expense_budgets"]
        budgets_coll.delete_many({"user_id": user_id})
        if budgets:
            for b in budgets:
                b["user_id"] = user_id
                budgets_coll.insert_one(b)
        return True

    @staticmethod
    def get_alerts(user_id):
        budgets = BudgetManager.get_budgets(user_id)
        alerts = []
        current_month = datetime.utcnow().strftime("%Y-%m")
        for b in budgets:
            cat = b.get("category", "")
            limit = float(b.get("limit", 0))
            pipeline = [
                {"$match": {"user_id": user_id, "type": "expense", "category": cat, "date": {"$regex": f"^{current_month}"}}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]
            result = list(Expense.get_collection().aggregate(pipeline))
            spent = result[0]["total"] if result else 0
            if spent > limit:
                alerts.append({
                    "category": cat,
                    "limit": limit,
                    "spent": spent,
                    "overshoot": round(spent - limit, 2),
                    "pct": round((spent / limit) * 100, 1) if limit > 0 else 0,
                })
        return alerts


class ExportService:
    @staticmethod
    def export_csv(user_id):
        import csv
        import io
        expenses = Expense.find_by_user(user_id)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Type", "Category", "Amount", "Description", "Date", "Created At"])
        for e in expenses:
            writer.writerow([e.type, e.category, e.amount, e.description, e.date, e.created_at])
        return output.getvalue()

    @staticmethod
    def export_excel(user_id):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Farm Finances"
        headers = ["Type", "Category", "Amount", "Description", "Date"]
        green_fill = PatternFill(start_color="1a7d36", end_color="1a7d36", fill_type="solid")
        white_font = Font(color="ffffff", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = green_fill
            cell.font = white_font
        expenses = Expense.find_by_user(user_id)
        for row, e in enumerate(expenses, 2):
            ws.cell(row=row, column=1, value=e.type)
            ws.cell(row=row, column=2, value=e.category)
            ws.cell(row=row, column=3, value=e.amount)
            ws.cell(row=row, column=4, value=e.description)
            ws.cell(row=row, column=5, value=e.date)
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_pdf(user_id):
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font("Arial", "", r"C:\Windows\Fonts\arial.ttf")
        pdf.add_font("Arial", "B", r"C:\Windows\Fonts\arialbd.ttf")
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, text="Farm Financial Report", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.set_font("Arial", "", 9)
        pdf.ln(8)
        summary = Expense.get_summary(user_id)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 7, text=f"Total Income: Rs. {summary.get('income', 0):.2f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 7, text=f"Total Expenses: Rs. {summary.get('expense', 0):.2f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 7, text=f"Net Profit: Rs. {summary.get('income', 0) - summary.get('expense', 0):.2f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(8)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 7, text="Transaction History", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font("Arial", "", 8)
        pdf.cell(30, 6, "Date", border=1)
        pdf.cell(25, 6, "Type", border=1)
        pdf.cell(40, 6, "Category", border=1)
        pdf.cell(20, 6, "Amount", border=1)
        pdf.cell(0, 6, "Description", border=1)
        pdf.ln()
        expenses = Expense.find_by_user(user_id)
        for e in expenses:
            pdf.cell(30, 5, e.date[:10], border=1)
            pdf.cell(25, 5, e.type, border=1)
            pdf.cell(40, 5, e.category[:25], border=1)
            pdf.cell(20, 5, f"Rs.{e.amount:.0f}", border=1)
            desc = (e.description or "")[:60]
            pdf.cell(0, 5, desc, border=1)
            pdf.ln()
        from io import BytesIO
        return BytesIO(bytes(pdf.output()))
