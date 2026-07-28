from flask import Blueprint, render_template, request, jsonify, session
from utils.auth import login_required
from models.crop import Crop
from utils.helpers import get_crop_list, get_districts, get_soil_types
from datetime import datetime
import traceback

crops_bp = Blueprint("crops", __name__)

SOIL_TYPES = [
    "Black Soil", "Red Soil", "Alluvial Soil", "Clay Soil",
    "Sandy Soil", "Loamy Soil", "Laterite Soil"
]

SEASONS = ["Kuruvai", "Samba", "Navarai", "Kharif", "Rabi", "Summer"]

STATUSES = ["Planned", "Seeded", "Growing", "Flowering", "Harvest Ready", "Harvested"]

DISTRICT_RECOMMENDATIONS = {
    "Thanjavur": [
        {"crop": "Paddy", "reason": "Staple crop of Cauvery Delta; high yield in alluvial soil", "soil": "Alluvial Soil", "season": "Samba", "duration": "120-150 days", "water": "High (800-1200mm)", "difficulty": "Moderate", "yield": "2.5-4 tons/acre"},
        {"crop": "Sugarcane", "reason": "Thrives in warm climate with clay-loam soil", "soil": "Clay Soil", "season": "Kuruvai", "duration": "10-12 months", "water": "Very High (1500-2500mm)", "difficulty": "High", "yield": "25-35 tons/acre"},
        {"crop": "Banana", "reason": "Ideal for delta region with good irrigation", "soil": "Alluvial Soil", "season": "Year-round", "duration": "10-12 months", "water": "High (1000-1500mm)", "difficulty": "Moderate", "yield": "12-18 tons/acre"},
        {"crop": "Black Gram", "reason": "Excellent rotation crop after paddy", "soil": "Clay Soil", "season": "Navarai", "duration": "70-90 days", "water": "Low (300-400mm)", "difficulty": "Easy", "yield": "0.5-0.8 tons/acre"},
    ],
    "Coimbatore": [
        {"crop": "Turmeric", "reason": "Premium quality turmeric; ideal climate and soil", "soil": "Red Soil", "season": "Kharif", "duration": "7-9 months", "water": "Moderate (600-900mm)", "difficulty": "Moderate", "yield": "2-3 tons/acre"},
        {"crop": "Maize", "reason": "Grows well in well-drained red soil", "soil": "Red Soil", "season": "Kharif", "duration": "85-110 days", "water": "Moderate (500-800mm)", "difficulty": "Easy", "yield": "1.5-2.5 tons/acre"},
        {"crop": "Coconut", "reason": "Perennial crop suited to Coimbatore climate", "soil": "Red Soil", "season": "Year-round", "duration": "Year-round", "water": "Moderate (600-1000mm)", "difficulty": "Moderate", "yield": "120-150 nuts/tree/year"},
        {"crop": "Groundnut", "reason": "Oilseed crop well-adapted to red loamy soil", "soil": "Loamy Soil", "season": "Rabi", "duration": "100-130 days", "water": "Moderate (400-600mm)", "difficulty": "Easy", "yield": "1-1.5 tons/acre"},
    ],
    "Salem": [
        {"crop": "Mango", "reason": "Premium variety mangoes; ideal climate for flowering", "soil": "Red Soil", "season": "Summer", "duration": "Perennial", "water": "Moderate (500-900mm)", "difficulty": "Moderate", "yield": "5-8 tons/acre"},
        {"crop": "Tapioca", "reason": "Staple root crop; high yield in red soil", "soil": "Red Soil", "season": "Kharif", "duration": "8-12 months", "water": "Low (300-500mm)", "difficulty": "Easy", "yield": "10-15 tons/acre"},
        {"crop": "Cotton", "reason": "Dryland cotton suited to Salem climate", "soil": "Black Soil", "season": "Kharif", "duration": "150-180 days", "water": "Moderate (500-700mm)", "difficulty": "Moderate", "yield": "1-1.5 tons/acre"},
        {"crop": "Groundnut", "reason": "Grown as rainfed crop in dry tracts", "soil": "Sandy Soil", "season": "Rabi", "duration": "100-130 days", "water": "Low (350-500mm)", "difficulty": "Easy", "yield": "0.8-1.2 tons/acre"},
    ],
    "Tiruppur": [
        {"crop": "Maize", "reason": "High demand for poultry feed industry", "soil": "Loamy Soil", "season": "Kharif", "duration": "85-110 days", "water": "Moderate (500-800mm)", "difficulty": "Easy", "yield": "2-3 tons/acre"},
        {"crop": "Groundnut", "reason": "Oilseed well-suited to dry conditions", "soil": "Sandy Soil", "season": "Rabi", "duration": "100-130 days", "water": "Low (350-500mm)", "difficulty": "Easy", "yield": "0.8-1.2 tons/acre"},
        {"crop": "Coconut", "reason": "Perennial crop suited to western zone", "soil": "Red Soil", "season": "Year-round", "duration": "Year-round", "water": "Moderate (600-1000mm)", "difficulty": "Moderate", "yield": "100-130 nuts/tree/year"},
        {"crop": "Millets", "reason": "Nutritious dryland crop; climate-resilient", "soil": "Sandy Soil", "season": "Kharif", "duration": "75-100 days", "water": "Low (250-400mm)", "difficulty": "Easy", "yield": "0.5-1 ton/acre"},
    ],
    "Cuddalore": [
        {"crop": "Paddy", "reason": "Major crop in Cuddalore with good irrigation", "soil": "Alluvial Soil", "season": "Samba", "duration": "120-150 days", "water": "High (800-1200mm)", "difficulty": "Moderate", "yield": "2.5-3.5 tons/acre"},
        {"crop": "Groundnut", "reason": "Well-suited to coastal sandy loam soil", "soil": "Sandy Soil", "season": "Navarai", "duration": "100-130 days", "water": "Low (350-500mm)", "difficulty": "Easy", "yield": "0.8-1.2 tons/acre"},
        {"crop": "Cashew", "reason": "Coastal climate ideal for cashew plantations", "soil": "Laterite Soil", "season": "Year-round", "duration": "Perennial", "water": "Low (300-500mm)", "difficulty": "Easy", "yield": "0.5-1 ton/acre"},
        {"crop": "Tapioca", "reason": "Root crop thrives in sandy loam", "soil": "Sandy Soil", "season": "Kharif", "duration": "8-12 months", "water": "Low (300-500mm)", "difficulty": "Easy", "yield": "8-12 tons/acre"},
    ],
    "Madurai": [
        {"crop": "Cotton", "reason": "Premium cotton grown in black cotton soil", "soil": "Black Soil", "season": "Kharif", "duration": "150-180 days", "water": "Moderate (500-700mm)", "difficulty": "Moderate", "yield": "1-1.5 tons/acre"},
        {"crop": "Paddy", "reason": "Grown in Vaigai delta region with irrigation", "soil": "Alluvial Soil", "season": "Samba", "duration": "120-150 days", "water": "High (800-1200mm)", "difficulty": "Moderate", "yield": "2-3.5 tons/acre"},
        {"crop": "Chilli", "reason": "Premium variety suited to dry climate", "soil": "Black Soil", "season": "Rabi", "duration": "120-150 days", "water": "Moderate (500-700mm)", "difficulty": "Moderate", "yield": "0.5-1 ton/acre"},
        {"crop": "Groundnut", "reason": "Grown as rainfed crop in dry areas", "soil": "Sandy Soil", "season": "Rabi", "duration": "100-130 days", "water": "Low (350-500mm)", "difficulty": "Easy", "yield": "0.8-1.2 tons/acre"},
    ],
    "Erode": [
        {"crop": "Turmeric", "reason": "High curcumin variety; Erode turmeric is GI-tagged", "soil": "Red Soil", "season": "Kharif", "duration": "7-9 months", "water": "Moderate (600-900mm)", "difficulty": "Moderate", "yield": "2.5-3.5 tons/acre"},
        {"crop": "Sugarcane", "reason": "Major sugarcane belt in Bhavani region", "soil": "Loamy Soil", "season": "Kuruvai", "duration": "10-12 months", "water": "Very High (1500-2500mm)", "difficulty": "High", "yield": "30-40 tons/acre"},
        {"crop": "Coconut", "reason": "Grown extensively in the river belt", "soil": "Alluvial Soil", "season": "Year-round", "duration": "Year-round", "water": "Moderate (600-1000mm)", "difficulty": "Moderate", "yield": "120-150 nuts/tree/year"},
        {"crop": "Banana", "reason": "Bhavani banana is premium quality", "soil": "Loamy Soil", "season": "Year-round", "duration": "10-12 months", "water": "High (1000-1500mm)", "difficulty": "Moderate", "yield": "15-20 tons/acre"},
    ],
    "Tirunelveli": [
        {"crop": "Paddy", "reason": "Grown in Tamirabarani delta region", "soil": "Alluvial Soil", "season": "Samba", "duration": "120-150 days", "water": "High (800-1200mm)", "difficulty": "Moderate", "yield": "2.5-3.5 tons/acre"},
        {"crop": "Coconut", "reason": "Coastal district ideal for coconut", "soil": "Sandy Soil", "season": "Year-round", "duration": "Year-round", "water": "Moderate (600-1000mm)", "difficulty": "Moderate", "yield": "100-130 nuts/tree/year"},
        {"crop": "Cotton", "reason": "Dryland cotton suited to rainfed areas", "soil": "Black Soil", "season": "Kharif", "duration": "150-180 days", "water": "Moderate (500-700mm)", "difficulty": "Moderate", "yield": "0.8-1.2 tons/acre"},
        {"crop": "Chilli", "reason": "Dry climate produces pungent chillies", "soil": "Red Soil", "season": "Rabi", "duration": "120-150 days", "water": "Moderate (500-700mm)", "difficulty": "Moderate", "yield": "0.5-1 ton/acre"},
    ],
    "Dharmapuri": [
        {"crop": "Mango", "reason": "Premium mango variety from Dharmapuri", "soil": "Red Soil", "season": "Summer", "duration": "Perennial", "water": "Low (400-700mm)", "difficulty": "Moderate", "yield": "5-7 tons/acre"},
        {"crop": "Tomato", "reason": "Major tomato-growing district in TN", "soil": "Red Soil", "season": "Rabi", "duration": "75-90 days", "water": "Moderate (400-600mm)", "difficulty": "Easy", "yield": "8-12 tons/acre"},
        {"crop": "Groundnut", "reason": "Dryland oilseed crop", "soil": "Sandy Soil", "season": "Kharif", "duration": "100-130 days", "water": "Low (350-500mm)", "difficulty": "Easy", "yield": "0.8-1.2 tons/acre"},
        {"crop": "Millets", "reason": "Climate-resilient nutri-cereals", "soil": "Red Soil", "season": "Kharif", "duration": "75-100 days", "water": "Low (250-400mm)", "difficulty": "Easy", "yield": "0.5-1 ton/acre"},
    ],
    "Virudhunagar": [
        {"crop": "Cotton", "reason": "Major cotton growing district", "soil": "Black Soil", "season": "Kharif", "duration": "150-180 days", "water": "Moderate (500-700mm)", "difficulty": "Moderate", "yield": "1-1.5 tons/acre"},
        {"crop": "Chilli", "reason": "Premium chilli variety for spice industry", "soil": "Red Soil", "season": "Rabi", "duration": "120-150 days", "water": "Moderate (500-700mm)", "difficulty": "Moderate", "yield": "0.5-1 ton/acre"},
        {"crop": "Groundnut", "reason": "Rainfed groundnut well-suited", "soil": "Sandy Soil", "season": "Kharif", "duration": "100-130 days", "water": "Low (350-500mm)", "difficulty": "Easy", "yield": "0.8-1.2 tons/acre"},
        {"crop": "Paddy", "reason": "Grown in limited irrigated areas", "soil": "Alluvial Soil", "season": "Samba", "duration": "120-150 days", "water": "High (800-1200mm)", "difficulty": "Moderate", "yield": "2-3 tons/acre"},
    ],
}

DEFAULT_RECOMMENDATIONS = [
    {"crop": "Paddy", "reason": "Staple food crop; suitable for most TN districts with irrigation", "soil": "Alluvial Soil", "season": "Samba", "duration": "120-150 days", "water": "High (800-1200mm)", "difficulty": "Moderate", "yield": "2.5-4 tons/acre"},
    {"crop": "Coconut", "reason": "Perennial cash crop suited to diverse soil types", "soil": "Sandy Soil", "season": "Year-round", "duration": "Year-round", "water": "Moderate (600-1000mm)", "difficulty": "Moderate", "yield": "100-150 nuts/tree/year"},
    {"crop": "Groundnut", "reason": "Popular oilseed with good market demand", "soil": "Sandy Soil", "season": "Rabi", "duration": "100-130 days", "water": "Low (350-500mm)", "difficulty": "Easy", "yield": "0.8-1.5 tons/acre"},
    {"crop": "Millets", "reason": "Nutritious; climate-resilient; low water requirement", "soil": "Red Soil", "season": "Kharif", "duration": "75-100 days", "water": "Low (250-400mm)", "difficulty": "Easy", "yield": "0.5-1 ton/acre"},
]

MONTH_NAMES_EN = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
MONTH_NAMES_TA = [
    "", "ஜனவரி", "பிப்ரவரி", "மார்ச்", "ஏப்ரல்", "மே", "ஜூன்",
    "ஜூலை", "ஆகஸ்ட்", "செப்டம்பர்", "அக்டோபர்", "நவம்பர்", "டிசம்பர்"
]


@crops_bp.route("/crops", methods=["GET"])
@login_required
def index():
    user_id = session.get("user_id")
    crops = Crop.find_by_user(user_id)
    lang = session.get("lang", "en")
    selected_district = session.get("district", "")
    stats = Crop.get_stats(user_id)
    activities = Crop.get_upcoming_activities(user_id) if user_id else []
    from utils.helpers import get_current_season
    current_season = get_current_season()
    now_year = datetime.utcnow().year
    current_month = datetime.utcnow().month
    return render_template(
        "crops.html",
        crops=[c.to_dict() for c in crops],
        crop_list=get_crop_list(),
        districts=get_districts(),
        soil_types=SOIL_TYPES,
        seasons=SEASONS,
        statuses=STATUSES,
        stats=stats,
        activities=activities,
        selected_district=selected_district,
        all_district_recommendations=DISTRICT_RECOMMENDATIONS,
        default_recommendations=DEFAULT_RECOMMENDATIONS,
        month_names_en=MONTH_NAMES_EN,
        month_names_ta=MONTH_NAMES_TA,
        current_season=current_season,
        now_year=now_year,
        current_month=current_month,
        lang=lang,
    )


@crops_bp.route("/api/crops", methods=["GET"])
@login_required
def get_crops():
    user_id = session.get("user_id")
    search_q = request.args.get("search", "").strip()
    filters = {}
    for key in ["district", "soil_type", "status", "season"]:
        val = request.args.get(key, "").strip()
        if val:
            filters[key] = val
    if search_q:
        crops = Crop.search_crops(user_id, search_q)
    elif filters:
        crops = Crop.filter_crops(user_id, filters)
    else:
        crops = Crop.find_by_user(user_id)
    return jsonify({"success": True, "crops": [c.to_dict() for c in crops]})


@crops_bp.route("/api/crops", methods=["POST"])
@login_required
def add_crop():
    try:
        data = request.get_json()
        user_id = session.get("user_id")
        lang = session.get("lang", "en")

        required = ["crop_name", "district", "soil_type", "land_size", "season", "planting_date", "harvest_date", "status"]
        for field in required:
            if not data.get(field):
                msg = f"{field} is required."
                return jsonify({"success": False, "message": msg})

        try:
            land_size = float(data["land_size"])
            if land_size <= 0:
                raise ValueError
        except (ValueError, TypeError):
            msg = "Land size must be a positive number." if lang == "en" else "நில அளவு நேர்மறை எண்ணாக இருக்க வேண்டும்."
            return jsonify({"success": False, "message": msg})

        crop = Crop()
        crop.user_id = user_id
        crop.crop_name = data["crop_name"]
        crop.village = data.get("village", "")
        crop.district = data["district"]
        crop.land_size = land_size
        crop.soil_type = data["soil_type"]
        crop.season = data["season"]
        crop.planting_date = data["planting_date"]
        crop.harvest_date = data["harvest_date"]
        crop.status = data["status"]
        crop.notes = data.get("notes", "")
        crop.save()

        stats = Crop.get_stats(user_id)
        msg = "Crop added successfully!" if lang == "en" else "பயிர் வெற்றிகரமாக சேர்க்கப்பட்டது!"
        return jsonify({"success": True, "message": msg, "stats": stats})
    except Exception as e:
        print(f"[Crop Error] add_crop: {traceback.format_exc()}")
        return jsonify({"success": False, "message": "Failed to add crop"}), 500


@crops_bp.route("/api/crops/<crop_id>", methods=["GET"])
@login_required
def get_crop(crop_id):
    try:
        lang = session.get("lang", "en")
        crop = Crop.find_by_id(crop_id)
        if not crop:
            msg = "Crop not found." if lang == "en" else "பயிர் கிடைக்கவில்லை."
            return jsonify({"success": False, "message": msg})
        return jsonify({"success": True, "crop": crop.to_dict()})
    except Exception as e:
        print(f"[Crop Error] get_crop: {traceback.format_exc()}")
        return jsonify({"success": False, "message": "Failed to load crop"}), 500


@crops_bp.route("/api/crops/<crop_id>", methods=["PUT"])
@login_required
def update_crop(crop_id):
    try:
        data = request.get_json()
        lang = session.get("lang", "en")

        crop = Crop.find_by_id(crop_id)
        if not crop:
            msg = "Crop not found." if lang == "en" else "பயிர் கிடைக்கவில்லை."
            return jsonify({"success": False, "message": msg})

        update_data = {}
        for field in ["crop_name", "district", "village", "soil_type", "season", "status", "planting_date", "harvest_date", "notes"]:
            if field in data:
                update_data[field] = data[field]
        if "land_size" in data:
            try:
                update_data["land_size"] = float(data["land_size"])
            except (ValueError, TypeError):
                msg = "Land size must be a number." if lang == "en" else "நில அளவு எண்ணாக இருக்க வேண்டும்."
                return jsonify({"success": False, "message": msg})

        crop.update(update_data)
        stats = Crop.get_stats(session.get("user_id"))
        msg = "Crop updated successfully!" if lang == "en" else "பயிர் வெற்றிகரமாக புதுப்பிக்கப்பட்டது!"
        return jsonify({"success": True, "message": msg, "stats": stats})
    except Exception as e:
        print(f"[Crop Error] update_crop: {traceback.format_exc()}")
        return jsonify({"success": False, "message": "Failed to update crop"}), 500


@crops_bp.route("/api/crops/<crop_id>", methods=["DELETE"])
@login_required
def delete_crop(crop_id):
    try:
        lang = session.get("lang", "en")
        crop = Crop.find_by_id(crop_id)
        if not crop:
            msg = "Crop not found." if lang == "en" else "பயிர் கிடைக்கவில்லை."
            return jsonify({"success": False, "message": msg})
        crop.delete()
        stats = Crop.get_stats(session.get("user_id"))
        msg = "Crop deleted successfully!" if lang == "en" else "பயிர் வெற்றிகரமாக நீக்கப்பட்டது!"
        return jsonify({"success": True, "message": msg, "stats": stats})
    except Exception as e:
        print(f"[Crop Error] delete_crop: {traceback.format_exc()}")
        return jsonify({"success": False, "message": "Failed to delete crop"}), 500


@crops_bp.route("/api/crops/stats", methods=["GET"])
@login_required
def get_stats():
    user_id = session.get("user_id")
    stats = Crop.get_stats(user_id)
    return jsonify({"success": True, "stats": stats})


@crops_bp.route("/api/crops/recommendations", methods=["GET"])
@login_required
def get_recommendations():
    district = request.args.get("district", "").strip()
    lang = session.get("lang", "en")
    recs = DISTRICT_RECOMMENDATIONS.get(district, DEFAULT_RECOMMENDATIONS)
    return jsonify({"success": True, "recommendations": recs})


@crops_bp.route("/api/crops/activities", methods=["GET"])
@login_required
def get_activities():
    user_id = session.get("user_id")
    activities = Crop.get_upcoming_activities(user_id)
    return jsonify({"success": True, "activities": activities})


@crops_bp.route("/api/crops/export", methods=["GET"])
@login_required
def export_crops():
    try:
        user_id = session.get("user_id")
        export_format = request.args.get("format", "csv")
        lang = session.get("lang", "en")
        crops = Crop.find_by_user(user_id)
        username = session.get("username", "Farmer")
        district = session.get("district", "")
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")

        if export_format == "csv":
            import io
            import csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["Crop Name", "Village", "District", "Land Size (Acres)", "Soil Type", "Season",
                             "Planting Date", "Expected Harvest", "Status", "Notes"])
            for c in crops:
                writer.writerow([c.crop_name, c.village, c.district, c.land_size, c.soil_type, c.season,
                                 c.planting_date, c.harvest_date, c.status, c.notes])
            csv_content = output.getvalue()
            output.close()
            return jsonify({
                "success": True,
                "export": csv_content,
                "filename": f"crops_export_{datetime.utcnow().strftime('%Y%m%d')}.csv",
                "mime": "text/csv",
            })
        elif export_format == "xlsx":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                return jsonify({"success": False, "message": "Excel export requires openpyxl. Install with: pip install openpyxl"})
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "My Crops"
            headers = ["Crop Name", "Village", "District", "Land Size (Acres)", "Soil Type", "Season",
                       "Planting Date", "Expected Harvest", "Status", "Notes"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            for row, c in enumerate(crops, 2):
                ws.cell(row=row, column=1, value=c.crop_name)
                ws.cell(row=row, column=2, value=c.village)
                ws.cell(row=row, column=3, value=c.district)
                ws.cell(row=row, column=4, value=c.land_size)
                ws.cell(row=row, column=5, value=c.soil_type)
                ws.cell(row=row, column=6, value=c.season)
                ws.cell(row=row, column=7, value=c.planting_date)
                ws.cell(row=row, column=8, value=c.harvest_date)
                ws.cell(row=row, column=9, value=c.status)
                ws.cell(row=row, column=10, value=c.notes)
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            import base64
            return jsonify({
                "success": True,
                "export": base64.b64encode(output.getvalue()).decode("ascii"),
                "filename": f"crops_export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "encoding": "base64",
            })
        else:
            from fpdf import FPDF
            from fpdf.enums import XPos, YPos
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_font("Arial", "", r"C:\Windows\Fonts\arial.ttf")
            pdf.add_font("Arial", "B", r"C:\Windows\Fonts\arialbd.ttf")
            pdf.set_font("Arial", "B", 16)
            label = "AI Agriculture Assistant - Crops Export" if lang == "en" else "AI விவசாய உதவியாளர் - பயிர் ஏற்றுமதி"
            pdf.cell(0, 10, text=label, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 6, text=f"Farmer: {username}  |  District: {district}  |  Date: {now}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2)
            pdf.ln(6)
            pdf.set_font("Arial", "B", 9)
            col_w = [32, 22, 22, 20, 22, 18, 18, 18, 16, 12]
            headers = ["Crop", "Village", "District", "Size(Ac)", "Soil", "Season", "Planted", "Harvest", "Status", "Notes"]
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 7, text=h, border=1, align="C")
            pdf.ln()
            pdf.set_font("Arial", "", 8)
            for c in crops:
                row_data = [c.crop_name[:12], c.village[:10], c.district[:10], str(c.land_size),
                            c.soil_type[:10], c.season[:8], c.planting_date[:8], c.harvest_date[:8],
                            c.status[:8], c.notes[:6]]
                for i, val in enumerate(row_data):
                    pdf.cell(col_w[i], 6, text=val, border=1, align="C")
                pdf.ln()
            output = bytes(pdf.output())
            import base64
            return jsonify({
                "success": True,
                "export": base64.b64encode(output).decode("ascii"),
                "filename": f"crops_export_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
                "mime": "application/pdf",
                "encoding": "base64",
            })
    except Exception as e:
        print(f"[Crop Error] export: {traceback.format_exc()}")
        return jsonify({"success": False, "message": "Failed to export crops"}), 500
